import openpyxl
import tkinter
from tkinter.filedialog import askopenfilename
from openpyxl.utils.exceptions import InvalidFileException


#==========================================================================================
# 
#                                     Cargar excel
#
#==========================================================================================
# Abre y ancla el archivo de excel a una variable.
#==========================================================================================

def cargar_excel(nombre_archivo):
	try:
		archivo_xlsx = openpyxl.load_workbook(nombre_archivo)
		print("Archivo ", nombre_archivo, " cargado.")

	except FileNotFoundError:
		print("No se encontró el archivo ", nombre_archivo)
		tkinter.Tk().withdraw()
		nombre_archivo = askopenfilename()
		archivo_xlsx = cargar_excel(nombre_archivo)
	
	except InvalidFileException:
		print("Archivo no valido, intenta de nuevo")
		tkinter.Tk().withdraw()
		nombre_archivo = askopenfilename()
		archivo_xlsx = cargar_excel(nombre_archivo)

	return archivo_xlsx




#==========================================================================================
# 
#                                     Cargar hoja
#
#==========================================================================================
# Abre y ancla una hoja del archivo excel a una variable.
#==========================================================================================

def cargar_hoja(archivo_excel, nombre_hoja):
	try:
		hoja_excel = archivo_excel.get_sheet_by_name(nombre_hoja)
		print("Hoja ", nombre_hoja, "cargada.")

	except KeyError:
		print("No se encontró la hoja ", nombre_hoja, " en el archivo excel")
		nombre_hoja = input("Ingersar el nombre de la hoja de excel: ")
		print(nombre_hoja)
		hoja_excel = cargar_hoja(archivo_excel, nombre_hoja)

	return hoja_excel




#==========================================================================================
# 
#                                 
#
#==========================================================================================
# 
#==========================================================================================

def actualizar_celdas(hoja, fila, columnas):
	celda = {}
	for nombre, columna in columnas.items():
		celda.update( {nombre : hoja.cell(row = fila, column = columna)} )
	return celda




#==========================================================================================
# 
#                                 
#
#==========================================================================================
# 
#==========================================================================================

def crear_hoja(archivo_excel, nombre_nueva_hoja):
	archivo_excel.create_sheet(nombre_nueva_hoja)
	hoja_nueva = cargar_hoja(archivo_excel, nombre_nueva_hoja)
	return hoja_nueva




#==========================================================================================
# 
#                                 
#
#==========================================================================================
# 
#==========================================================================================

def copiar_rango(hoja, fila_ini, col_ini, fila_fin, col_fin):
	rango_completo = []
	rango_fila = []

	for fila in range(fila_ini, fila_fin+1):
		for columna in range(col_ini, col_fin+1):
			rango_fila.append( hoja.cell(row=fila, column=columna).value )
		rango_completo.append( rango_fila )

	return rango_completo





#==========================================================================================
# 
#                                 
#
#==========================================================================================
# 
#==========================================================================================

def pegar_rango(hoja, lista, fila_inicial, columna_inicial):
	
	fila_actual = fila_inicial
	columna_actual = columna_inicial

	for fila in lista:
		
		for celda in fila:
			hoja.cell(row=fila_actual, column=columna_actual).value = celda
			columna_actual = columna_actual + 1

		fila_actual	= fila_actual + 1
		columna_actual = columna_inicial		





#==========================================================================================
# 
#                                 
#
#==========================================================================================
# 
#==========================================================================================

def copiar_encabezado(hoja_con_encabezado, hoja_sin_encabezado):
	lista_encabezado = copiar_rango(hoja_con_encabezado, 1, 1, 1, 100)
	pegar_rango(hoja_sin_encabezado, lista_encabezado, 1, 1)






#==========================================================================================
# 
#                                 
#
#==========================================================================================
# 
#==========================================================================================

def copiar_fila(hoja, fila):
	lista_fila = copiar_rango(hoja, fila, 1, fila, 100)
	return lista_fila





#==========================================================================================
# 
#                                 
#
#==========================================================================================
# 
#==========================================================================================

def pegar_fila(hoja, lista_fila, fila):
	pegar_rango(hoja, lista_fila, fila, 1)






#==========================================================================================
# 
#                                 
#
#==========================================================================================
# 
#==========================================================================================

def guardar_xlsx(xlsx_archivo, nombre, intento=0):
	try:
		if intento > 0:
			nombre_archivo = nombre + "_" + str(intento) + ".xlsx"
		else:
			nombre_archivo = nombre + ".xlsx"
		
		xlsx_archivo.save(nombre_archivo)
		print("Archivo guardado como ", nombre_archivo)

	except PermissionError:
		guardar_xlsx(xlsx_archivo, nombre, intento + 1)
		

	


